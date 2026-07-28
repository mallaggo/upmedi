from difflib import get_close_matches
from blog.models import MyProduct
from django.db.models import Value
from django.db.models.functions import Replace
import os
import openpyxl
from django.conf import settings


SEARCH_PRODUCT_TOOL = {
    "type": "function",
    "name": "search_product",
    "description": (
    "우리 쇼핑몰 데이터베이스(MyProduct)에 등록된 상품을 검색한다. "
    "사용자가 상품명, 카테고리, 가격, 재고 등 상품을 찾거나 조회하는 요청을 하면 반드시 이 Tool을 호출한다. "
    "등록되지 않은 상품을 추측하거나 일반적인 상품 추천을 하지 않는다."
    ),
    "parameters": {
        "type": "object",
        "properties": {

            "keyword": {
                "type": "string",
                "description": "상품명"
            },

            "category": {
                "type": "string",
                "description": "카테고리명"
            },

            "min_price": {
                "type": "integer",
                "description": "최소 가격"
            },

            "max_price": {
                "type": "integer",
                "description": "최대 가격"
            },
            "stock": {
            "type": "integer",
            "description": "재고가 이 값 이하인 상품 검색"
            }

        }
    }
}

READ_EXCEL_PRODUCTS_TOOL = {
    "type": "function",
    "name": "read_excel_products",
    "description": "현재 업로드된 엑셀 파일에서 상품 정보를 읽는다.",
    "parameters": {
        "type": "object",
        "properties": {},
        "additionalProperties": False
    }
}



def search_product(keyword=None, category=None, min_price=None, max_price=None, stock=None, context=None):

    products = MyProduct.objects.select_related("category")

    # 기본값은 사용자가 입력한 검색어
    real_name = keyword

    if keyword:
        products = products.annotate(
            clean_name=Replace("name", Value(" "), Value(""))
        ).filter(
            clean_name__icontains=keyword.replace(" ", "")
        )

        # 정확히 찾지 못하면 유사도 검색
        if not products.exists():

            names = list(
                MyProduct.objects.values_list("name", flat=True)
            )

            match = get_close_matches(
                keyword.replace(" ", ""),
                [name.replace(" ", "") for name in names],
                n=1,
                cutoff=0.5
            )



            if match:
                for name in names:
                    if name.replace(" ", "") == match[0]:
                        real_name = name
                        break

                products = MyProduct.objects.select_related(
                    "category"
                ).filter(
                    name=real_name
                )

    if category:
        products = products.filter(category__name__icontains=category)

    if min_price is not None:
        products = products.filter(price__gte=min_price)

    if max_price is not None:
        products = products.filter(price__lte=max_price)

    if stock is not None:
        products = products.filter(stock__lte=stock)

    result = []

    for p in products:
        result.append({
            "id": p.id,
            "name": p.name,
            "category": p.category.name,
            "price": p.price,
            "stock": p.stock,
        })

    if not result:
        return {
            "success": False,
            "message": "등록된 상품이 없습니다.",
            "corrected_keyword": real_name,
            "data": []
        }

    return {
        "success": True,
        "message": f"{len(result)}개의 상품을 찾았습니다.",
        "corrected_keyword": real_name,
        "data": result
    }




def read_excel_products(context=None):
    """
    엑셀 파일을 읽어 상품 데이터를 반환한다.
    DB에는 저장하지 않는다.   """
    filename = context.get("uploaded_filename")

    filepath = os.path.join(
        settings.MEDIA_ROOT,
        "excel",
        filename
    )

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # -----------------------------
    # 헤더 읽기
    # -----------------------------
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

    header_map = {
        "카테고리": "category",
        "상품명": "name",
        "가격": "price",
        "재고": "stock",
        "설명": "short_desc",
    }

    columns = {}

    for idx, header in enumerate(headers):
        if header in header_map:
            columns[header_map[header]] = idx

    # 필수 컬럼 검사
    required = ["name", "price"]

    missing = [field for field in required if field not in columns]

    if missing:
        return {
            "success": False,
            "message": f"필수 컬럼이 없습니다. ({', '.join(missing)})"
        }

    products = []
    errors = []

    # -----------------------------
    # 데이터 읽기
    # -----------------------------
    for row_num, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True),
            start=2
    ):

        if not any(row):
            continue

        category = row[columns["category"]] if "category" in columns else ""
        name = row[columns["name"]]
        price = row[columns["price"]]
        stock = row[columns["stock"]] if "stock" in columns else 0
        short_desc = row[columns["short_desc"]] if "short_desc" in columns else ""

        # 상품명
        if not name:
            errors.append({
                "row": row_num,
                "message": "상품명이 없습니다."
            })
            continue

        # 가격
        try:
            price = int(price or 0)
        except:
            errors.append({
                "row": row_num,
                "message": "가격이 숫자가 아닙니다."
            })
            continue

        # 재고
        try:
            stock = int(stock or 0)
        except:
            errors.append({
                "row": row_num,
                "message": "재고가 숫자가 아닙니다."
            })
            continue

        products.append({
            "row": row_num,
            "category": str(category or "").strip(),
            "name": str(name).strip(),
            "price": price,
            "stock": stock,
            "short_desc": str(short_desc or "").strip(),
        })

    result = {
        "success": len(errors) == 0,
        "message": f"{len(products)}개의 상품을 읽었습니다.",
        "data": products,
        "errors": errors,
    }

    if context is not None:
        context.set("read_excel_products", result)

    return result